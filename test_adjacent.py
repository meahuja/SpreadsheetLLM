"""Create a small Excel with 2 adjacent tables (no gap), encode it,
and verify the output step by step."""
import json
import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.path.insert(0, os.path.dirname(__file__))
from spreadsheet_llm.encoder import encode_spreadsheet
from spreadsheet_llm.vanilla import vanilla_encode

# ==========================================================================
# CREATE THE EXCEL
# ==========================================================================
path = os.path.join(os.path.dirname(__file__), "adjacent_2tables.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"

BOLD = Font(bold=True)
BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GREEN_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(bottom=Side(style="thin"))

# ------ TABLE 1: Product Sales (rows 1-8) ------
# Header row 1
headers1 = ["Product", "Category", "Q1 Sales", "Q2 Sales", "Total", "Growth%"]
for c, h in enumerate(headers1, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = WHITE_FONT
    cell.fill = BLUE_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER

# Data rows 2-7
products = [
    ("Laptop Pro",    "Electronics", 15000, 18500),
    ("Wireless Mouse","Electronics",  8200,  9100),
    ("Office Chair",  "Furniture",   12000, 11000),
    ("Standing Desk", "Furniture",   22000, 28000),
    ("USB-C Hub",     "Accessories",  5500,  7200),
    ("Monitor 27in",  "Electronics", 31000, 35000),
]
for i, (prod, cat, q1, q2) in enumerate(products):
    r = i + 2
    ws.cell(row=r, column=1, value=prod)
    ws.cell(row=r, column=2, value=cat)
    ws.cell(row=r, column=3, value=q1).number_format = '"$"#,##0'
    ws.cell(row=r, column=4, value=q2).number_format = '"$"#,##0'
    ws.cell(row=r, column=5, value=f"=C{r}+D{r}").number_format = '"$"#,##0'
    ws.cell(row=r, column=6, value=f"=(D{r}-C{r})/C{r}").number_format = "0.0%"

# Total row 8
ws.cell(row=8, column=1, value="TOTAL").font = BOLD
ws.cell(row=8, column=3, value="=SUM(C2:C7)").number_format = '"$"#,##0'
ws.cell(row=8, column=4, value="=SUM(D2:D7)").number_format = '"$"#,##0'
ws.cell(row=8, column=5, value="=SUM(E2:E7)").number_format = '"$"#,##0'
ws.cell(row=8, column=6, value="=(D8-C8)/C8").number_format = "0.0%"

# ------ TABLE 2: Employee Directory (rows 9-16) — NO GAP! ------
headers2 = ["Name", "Role", "Department", "Hire Date", "Salary", "Email"]
for c, h in enumerate(headers2, 1):
    cell = ws.cell(row=9, column=c, value=h)
    cell.font = WHITE_FONT
    cell.fill = GREEN_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER

employees = [
    ("Alice Johnson",  "Engineer",    "Engineering", date(2021, 3, 15), 95000, "alice@company.com"),
    ("Bob Williams",   "Manager",     "Sales",       date(2019, 7, 1),  88000, "bob@company.com"),
    ("Carol Davis",    "Designer",    "Marketing",   date(2022, 1, 10), 82000, "carol@company.com"),
    ("Dave Martinez",  "Analyst",     "Finance",     date(2020, 9, 20), 78000, "dave@company.com"),
    ("Eve Thompson",   "Engineer",    "Engineering", date(2023, 6, 5),  91000, "eve@company.com"),
    ("Frank Wilson",   "Director",    "Sales",       date(2018, 2, 14), 115000,"frank@company.com"),
    ("Grace Lee",      "Engineer",    "Engineering", date(2021, 11, 8), 97000, "grace@company.com"),
]
for i, (name, role, dept, hire, salary, email) in enumerate(employees):
    r = 10 + i
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=role)
    ws.cell(row=r, column=3, value=dept)
    ws.cell(row=r, column=4, value=hire).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=5, value=salary).number_format = '"$"#,##0'
    ws.cell(row=r, column=6, value=email)

wb.save(path)
print(f"Created: {path}")
print(f"  Table 1: Product Sales  (rows 1-8,  cols A-F)")
print(f"  Table 2: Employee Dir   (rows 9-16, cols A-F)")
print(f"  NO empty row between them!")
print()

# ==========================================================================
# SHOW RAW EXCEL CONTENT
# ==========================================================================
print("="*70)
print("RAW EXCEL CONTENT (what's in each cell)")
print("="*70)
wb2 = openpyxl.load_workbook(path, data_only=False)
ws2 = wb2.active
for r in range(1, ws2.max_row + 1):
    row_vals = []
    for c in range(1, ws2.max_column + 1):
        v = ws2.cell(row=r, column=c).value
        if v is None:
            row_vals.append("")
        else:
            row_vals.append(str(v)[:20])
    marker = ""
    if r == 1: marker = " <-- TABLE 1 HEADER"
    elif r == 8: marker = " <-- TABLE 1 TOTAL"
    elif r == 9: marker = " <-- TABLE 2 HEADER (no gap!)"
    elif r == 16: marker = " <-- TABLE 2 LAST ROW"
    print(f"  Row {r:2d}: {' | '.join(f'{v:20s}' for v in row_vals)}{marker}")

# ==========================================================================
# ENCODE
# ==========================================================================
print()
print("="*70)
print("ENCODING WITH SHEETCOMPRESSOR")
print("="*70)

result = encode_spreadsheet(path, k=2)
sheet = result["sheets"]["Data"]

# Anchors
anchors = sheet["structural_anchors"]
print(f"\nStructural Anchors:")
print(f"  Rows: {anchors['rows']}")
print(f"  Cols: {anchors['columns']}")

# Check table boundary detection
row_set = set(anchors["rows"])
t1_header = 1 in row_set or 2 in row_set
t2_header = any(r in row_set for r in [8, 9, 10])
print(f"\n  Table 1 header (row 1) detected: {'YES' if t1_header else 'NO'}")
print(f"  Table 2 header (row 9) detected: {'YES' if t2_header else 'NO'}")

# Cells
cells = sheet["cells"]
print(f"\nCells (inverted index): {len(cells)} unique values")

print(f"\n  Formulas:")
for val, refs in cells.items():
    if val.startswith("="):
        print(f"    {val:35s} -> {refs}")

print(f"\n  Table 1 values:")
for val, refs in cells.items():
    if not val.startswith("=") and any("company.com" not in val and r in ["A1", "B1", "C1", "A2", "B2"] for r in refs):
        pass  # skip

t1_values = ["Laptop Pro", "Wireless Mouse", "Office Chair", "Standing Desk",
             "USB-C Hub", "Monitor 27in", "Electronics", "Furniture", "Accessories", "TOTAL"]
t2_values = ["Alice Johnson", "Bob Williams", "Carol Davis", "Dave Martinez",
             "Eve Thompson", "Frank Wilson", "Grace Lee", "Engineer", "Manager",
             "Designer", "Analyst", "Director"]

print(f"\n  Table 1 products found in cells:")
for v in t1_values:
    if v in cells:
        print(f"    {v:25s} -> {cells[v]}")

print(f"\n  Table 2 employees found in cells:")
for v in t2_values:
    if v in cells:
        print(f"    {v:25s} -> {cells[v]}")

# Formats
fmts = sheet["formats"]
print(f"\nFormats ({len(fmts)} groups):")
for fmt_key, refs in fmts.items():
    parsed = json.loads(fmt_key)
    t = parsed["type"]
    nfs = parsed["nfs"]
    print(f"  {t:15s} nfs={nfs:25s} -> {refs}")

# Metrics
m = result["compression_metrics"]["overall"]
print(f"\nCompression Metrics:")
print(f"  Original:    {m['original_tokens']:>6,} tokens")
print(f"  Final:       {m['final_tokens']:>6,} tokens")
print(f"  Ratio:       {m['overall_ratio']:.2f}x")
print(f"  Format agg:  {m['format_ratio']:.2f}x")

# Vanilla comparison
print(f"\n{'='*70}")
print("VANILLA ENCODING (for comparison)")
print("="*70)
vanilla = vanilla_encode(path)
print(vanilla["Data"])

# Full JSON
print(f"\n{'='*70}")
print("FULL COMPRESSED JSON")
print("="*70)
print(json.dumps(result, indent=2, ensure_ascii=False))
