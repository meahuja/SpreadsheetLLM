"""Comprehensive demo & test for SpreadsheetLLM encoder.

Creates test spreadsheets of varying complexity and size, runs encoding,
verifies correctness of output including formula preservation.
"""

from __future__ import annotations

import json
import logging
import os
import sys
import time
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Add project root to path
sys.path.insert(0, os.path.dirname(__file__))

from spreadsheet_llm.encoder import encode_spreadsheet
from spreadsheet_llm.vanilla import vanilla_encode

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    bottom=Side(style="thin"),
    top=Side(style="thin"),
    left=Side(style="thin"),
    right=Side(style="thin"),
)
CURRENCY_FMT = '"$"#,##0.00'
PCT_FMT = "0.00%"
DATE_FMT = "yyyy-mm-dd"
INT_FMT = "#,##0"


# =============================================================================
# Test spreadsheet generators
# =============================================================================


def create_tiny_sheet(path: str) -> None:
    """Tiny: 1 table, 5 rows, basic data. Tests minimum viable input."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = ["Product", "Q1", "Q2", "Total"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = BOLD
        cell.alignment = CENTER

    data = [
        ("Widget A", 100, 200, "=B2+C2"),
        ("Widget B", 150, 300, "=B3+C3"),
        ("Widget C", 75, 125, "=B4+C4"),
        ("Total", "=SUM(B2:B4)", "=SUM(C2:C4)", "=SUM(D2:D4)"),
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 2 and r < 5:
                cell.number_format = INT_FMT

    wb.save(path)
    logger.info("Created tiny sheet: %s (5 rows x 4 cols)", path)


def create_small_sheet(path: str) -> None:
    """Small: 2 tables on same sheet (adjacent, no gap), mixed types."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Table 1: Sales data (rows 1-8) ---
    headers1 = ["Region", "Revenue", "Cost", "Profit", "Margin"]
    for c, h in enumerate(headers1, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    regions = ["North", "South", "East", "West", "Central"]
    for i, region in enumerate(regions):
        r = i + 2
        ws.cell(row=r, column=1, value=region)
        ws.cell(row=r, column=2, value=(i + 1) * 10000).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=(i + 1) * 6000).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = CURRENCY_FMT
        ws.cell(row=r, column=5, value=f"=D{r}/B{r}").number_format = PCT_FMT

    # Totals row
    r = 7
    ws.cell(row=r, column=1, value="Total").font = BOLD
    ws.cell(row=r, column=2, value="=SUM(B2:B6)").number_format = CURRENCY_FMT
    ws.cell(row=r, column=3, value="=SUM(C2:C6)").number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value="=SUM(D2:D6)").number_format = CURRENCY_FMT
    ws.cell(row=r, column=5, value="=D7/B7").number_format = PCT_FMT

    # --- Table 2: Employee data (rows 8-14, NO gap row!) ---
    headers2 = ["Name", "Department", "Start Date", "Salary", "Email"]
    for c, h in enumerate(headers2, 1):
        cell = ws.cell(row=8, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    employees = [
        ("Alice Smith", "Engineering", date(2020, 3, 15), 95000, "alice@company.com"),
        ("Bob Jones", "Marketing", date(2019, 7, 1), 82000, "bob@company.com"),
        ("Carol White", "Engineering", date(2021, 1, 10), 105000, "carol@company.com"),
        ("Dave Brown", "Sales", date(2018, 11, 20), 78000, "dave@company.com"),
        ("Eve Black", "Engineering", date(2022, 6, 5), 112000, "eve@company.com"),
    ]
    for i, emp in enumerate(employees):
        r = 9 + i
        ws.cell(row=r, column=1, value=emp[0])
        ws.cell(row=r, column=2, value=emp[1])
        ws.cell(row=r, column=3, value=emp[2]).number_format = DATE_FMT
        ws.cell(row=r, column=4, value=emp[3]).number_format = CURRENCY_FMT
        ws.cell(row=r, column=5, value=emp[4])

    # Average salary formula
    ws.cell(row=14, column=1, value="Average Salary").font = BOLD
    ws.cell(row=14, column=4, value="=AVERAGE(D9:D13)").number_format = CURRENCY_FMT

    # --- Sheet 2: Summary with merged header ---
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:D1")
    ws2.cell(row=1, column=1, value="Annual Summary Report").font = Font(bold=True, size=14)
    ws2.cell(row=1, column=1).alignment = CENTER

    ws2.cell(row=3, column=1, value="Metric").font = BOLD
    ws2.cell(row=3, column=2, value="Value").font = BOLD
    ws2.cell(row=4, column=1, value="Total Revenue")
    ws2.cell(row=4, column=2, value="=Report!B7")
    ws2.cell(row=5, column=1, value="Total Cost")
    ws2.cell(row=5, column=2, value="=Report!C7")
    ws2.cell(row=6, column=1, value="Net Profit")
    ws2.cell(row=6, column=2, value="=Report!D7")
    ws2.cell(row=7, column=1, value="Headcount")
    ws2.cell(row=7, column=2, value="=COUNTA(Report!A9:A13)")

    wb.save(path)
    logger.info("Created small sheet: %s (2 tables, 2 sheets, merged header)", path)


def create_medium_sheet(path: str) -> None:
    """Medium: 500 rows, multiple data types, complex formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "ID", "Date", "Customer", "Product", "Qty", "Unit Price",
        "Subtotal", "Tax Rate", "Tax", "Total", "Status",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Service Z"]
    customers = ["Acme Corp", "Globex Inc", "Initech", "Umbrella Co", "Stark Ind"]
    statuses = ["Completed", "Pending", "Shipped", "Processing", "Completed"]

    for r in range(2, 502):
        idx = r - 2
        ws.cell(row=r, column=1, value=f"TXN-{idx + 1:04d}")
        ws.cell(row=r, column=2, value=date(2024, (idx % 12) + 1, (idx % 28) + 1)).number_format = DATE_FMT
        ws.cell(row=r, column=3, value=customers[idx % 5])
        ws.cell(row=r, column=4, value=products[idx % 5])
        ws.cell(row=r, column=5, value=(idx % 10) + 1).number_format = INT_FMT
        ws.cell(row=r, column=6, value=round(10.0 + (idx % 50) * 2.5, 2)).number_format = CURRENCY_FMT
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}").number_format = CURRENCY_FMT
        ws.cell(row=r, column=8, value=0.08 if idx % 3 == 0 else 0.10).number_format = PCT_FMT
        ws.cell(row=r, column=9, value=f"=G{r}*H{r}").number_format = CURRENCY_FMT
        ws.cell(row=r, column=10, value=f"=G{r}+I{r}").number_format = CURRENCY_FMT
        ws.cell(row=r, column=11, value=statuses[idx % 5])

    # Summary row
    r = 502
    ws.cell(row=r, column=1, value="TOTALS").font = BOLD
    ws.cell(row=r, column=5, value="=SUM(E2:E501)").number_format = INT_FMT
    ws.cell(row=r, column=7, value="=SUM(G2:G501)").number_format = CURRENCY_FMT
    ws.cell(row=r, column=9, value="=SUM(I2:I501)").number_format = CURRENCY_FMT
    ws.cell(row=r, column=10, value="=SUM(J2:J501)").number_format = CURRENCY_FMT

    # Stats formulas
    r = 504
    ws.cell(row=r, column=1, value="Statistics").font = Font(bold=True, size=12)
    ws.cell(row=505, column=1, value="Avg Order Value")
    ws.cell(row=505, column=2, value="=AVERAGE(J2:J501)").number_format = CURRENCY_FMT
    ws.cell(row=506, column=1, value="Max Order")
    ws.cell(row=506, column=2, value="=MAX(J2:J501)").number_format = CURRENCY_FMT
    ws.cell(row=507, column=1, value="Min Order")
    ws.cell(row=507, column=2, value="=MIN(J2:J501)").number_format = CURRENCY_FMT
    ws.cell(row=508, column=1, value="Unique Customers")
    ws.cell(row=508, column=2, value="=COUNTA(C2:C501)/COUNTIF(C2:C501,C2)")
    ws.cell(row=509, column=1, value="Completed %")
    ws.cell(row=509, column=2, value='=COUNTIF(K2:K501,"Completed")/COUNTA(K2:K501)').number_format = PCT_FMT

    wb.save(path)
    logger.info("Created medium sheet: %s (502 rows x 11 cols, complex formulas)", path)


def create_large_sheet(path: str) -> None:
    """Large: 5000 rows, 3 tables on one sheet + 2 extra sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Table 1: rows 1-2001 (2000 data rows) ---
    headers1 = ["ID", "Timestamp", "Sensor", "Reading", "Unit", "Status", "Delta"]
    for c, h in enumerate(headers1, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    sensors = ["TEMP-01", "TEMP-02", "HUM-01", "PRESS-01", "FLOW-01"]
    units = ["°C", "°C", "%RH", "kPa", "L/min"]

    for r in range(2, 2002):
        idx = r - 2
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=datetime(2024, 1, 1, idx % 24, (idx * 5) % 60)).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=r, column=3, value=sensors[idx % 5])
        ws.cell(row=r, column=4, value=round(20.0 + (idx % 100) * 0.3, 2))
        ws.cell(row=r, column=5, value=units[idx % 5])
        ws.cell(row=r, column=6, value="OK" if idx % 7 != 0 else "ALERT")
        if r > 2:
            ws.cell(row=r, column=7, value=f"=D{r}-D{r-1}")

    # --- Table 2: rows 2003-2503 (500 rows, adjacent after 1 empty row) ---
    ws.cell(row=2003, column=1, value="Location").font = BOLD
    ws.cell(row=2003, column=2, value="Avg Temp").font = BOLD
    ws.cell(row=2003, column=3, value="Max Temp").font = BOLD
    ws.cell(row=2003, column=4, value="Alerts").font = BOLD

    locations = ["Building A", "Building B", "Building C", "Warehouse", "Lab"]
    for r in range(2004, 2504):
        idx = r - 2004
        ws.cell(row=r, column=1, value=locations[idx % 5])
        ws.cell(row=r, column=2, value=round(21.0 + (idx % 10) * 0.5, 1))
        ws.cell(row=r, column=3, value=round(25.0 + (idx % 15) * 0.3, 1))
        ws.cell(row=r, column=4, value=idx % 3)

    # --- Table 3: rows 2504-3004 (NO gap! Adjacent to table 2) ---
    headers3 = ["Shift", "Operator", "Start", "End", "Output", "Efficiency"]
    for c, h in enumerate(headers3, 1):
        cell = ws.cell(row=2504, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    shifts = ["Morning", "Afternoon", "Night"]
    operators = ["John", "Jane", "Bob", "Alice", "Tom"]
    for r in range(2505, 3005):
        idx = r - 2505
        ws.cell(row=r, column=1, value=shifts[idx % 3])
        ws.cell(row=r, column=2, value=operators[idx % 5])
        start_hour = (6 + (idx % 3) * 8) % 24
        end_hour = (start_hour + 8) % 24
        ws.cell(row=r, column=3, value=datetime(2024, 1, 1, start_hour, 0)).number_format = "hh:mm"
        ws.cell(row=r, column=4, value=datetime(2024, 1, 1, end_hour, 0)).number_format = "hh:mm"
        ws.cell(row=r, column=5, value=50 + idx % 30).number_format = INT_FMT
        ws.cell(row=r, column=6, value=f"=E{r}/80").number_format = PCT_FMT

    # --- Remaining rows to ~5000 ---
    for r in range(3006, 5001):
        ws.cell(row=r, column=1, value=f"PAD-{r}")

    # --- Sheet 2: Pivot-like summary ---
    ws2 = wb.create_sheet("Pivot")
    ws2.merge_cells("A1:E1")
    ws2.cell(row=1, column=1, value="Sensor Summary").font = Font(bold=True, size=14)

    ws2.cell(row=3, column=1, value="Sensor").font = BOLD
    ws2.cell(row=3, column=2, value="Count").font = BOLD
    ws2.cell(row=3, column=3, value="Avg").font = BOLD
    ws2.cell(row=3, column=4, value="Min").font = BOLD
    ws2.cell(row=3, column=5, value="Max").font = BOLD

    for i, sensor in enumerate(sensors):
        r = 4 + i
        ws2.cell(row=r, column=1, value=sensor)
        ws2.cell(row=r, column=2, value=f'=COUNTIF(Data!C:C,"{sensor}")')
        ws2.cell(row=r, column=3, value=f'=AVERAGEIF(Data!C:C,"{sensor}",Data!D:D)')
        ws2.cell(row=r, column=4, value=f'=MINIFS(Data!D:D,Data!C:C,"{sensor}")')
        ws2.cell(row=r, column=5, value=f'=MAXIFS(Data!D:D,Data!C:C,"{sensor}")')

    # --- Sheet 3: Config (form-like, sparse) ---
    ws3 = wb.create_sheet("Config")
    ws3.cell(row=1, column=1, value="Configuration").font = Font(bold=True, size=14)
    config_items = [
        ("Sample Rate (Hz)", 10),
        ("Alert Threshold (°C)", 35.0),
        ("Data Retention (days)", 365),
        ("Email Notifications", True),
        ("Admin Email", "admin@factory.com"),
    ]
    for i, (label, val) in enumerate(config_items):
        ws3.cell(row=3 + i, column=1, value=label).font = BOLD
        ws3.cell(row=3 + i, column=3, value=val)

    wb.save(path)
    logger.info("Created large sheet: %s (5000 rows, 3 tables, 3 sheets)", path)


# =============================================================================
# Verification
# =============================================================================


def verify_formula_preservation(encoding: dict, label: str) -> bool:
    """Check that formulas (=...) appear in the encoded cells, not computed values."""
    formula_found = False
    for sheet_name, sheet_data in encoding.get("sheets", {}).items():
        cells = sheet_data.get("cells", {})
        for value, refs in cells.items():
            if str(value).startswith("="):
                formula_found = True
                break
        if formula_found:
            break

    if formula_found:
        print(f"  [PASS] {label}: Formulas preserved in encoding")
    else:
        print(f"  [WARN] {label}: No formulas found in encoding (may have no formula cells)")
    return formula_found


def verify_structure(encoding: dict, label: str) -> bool:
    """Verify the encoding has the expected structure."""
    ok = True
    for sheet_name, sheet_data in encoding.get("sheets", {}).items():
        if "structural_anchors" not in sheet_data:
            print(f"  [FAIL] {label}/{sheet_name}: missing structural_anchors")
            ok = False
        else:
            anchors = sheet_data["structural_anchors"]
            if not anchors.get("rows") or not anchors.get("columns"):
                print(f"  [WARN] {label}/{sheet_name}: empty anchors (rows={len(anchors.get('rows', []))}, cols={len(anchors.get('columns', []))})")

        if "cells" not in sheet_data:
            print(f"  [FAIL] {label}/{sheet_name}: missing cells")
            ok = False

        if "formats" not in sheet_data:
            print(f"  [FAIL] {label}/{sheet_name}: missing formats")
            ok = False

    if ok:
        print(f"  [PASS] {label}: Structure valid")
    return ok


def verify_ranges_merged(encoding: dict, label: str) -> bool:
    """Check that inverted index uses ranges (A1:A5) not just single refs."""
    range_found = False
    for sheet_data in encoding.get("sheets", {}).values():
        for value, refs in sheet_data.get("cells", {}).items():
            for ref in refs:
                if ":" in ref:
                    range_found = True
                    break
            if range_found:
                break
        if range_found:
            break

    if range_found:
        print(f"  [PASS] {label}: Cell ranges properly merged")
    else:
        print(f"  [INFO] {label}: No merged ranges found (may be too small)")
    return range_found


def print_metrics(encoding: dict, label: str) -> None:
    """Print compression metrics."""
    metrics = encoding.get("compression_metrics", {}).get("overall", {})
    orig = metrics.get("original_tokens", 0)
    final = metrics.get("final_tokens", 0)
    ratio = metrics.get("overall_ratio", 0)

    print(f"\n  {label} Compression Metrics:")
    print(f"    Original tokens:       {orig:>10,}")
    print(f"    After anchors:         {metrics.get('after_anchor_tokens', 0):>10,}  ({metrics.get('anchor_ratio', 0):.2f}x)")
    print(f"    After inverted index:  {metrics.get('after_inverted_index_tokens', 0):>10,}  ({metrics.get('inverted_index_ratio', 0):.2f}x)")
    print(f"    After format agg:      {metrics.get('after_format_tokens', 0):>10,}  ({metrics.get('format_ratio', 0):.2f}x)")
    print(f"    Final:                 {final:>10,}  ({ratio:.2f}x)")


def print_sample_formulas(encoding: dict) -> None:
    """Print a few sample formula entries from the encoding."""
    formulas = []
    for sheet_data in encoding.get("sheets", {}).values():
        for value, refs in sheet_data.get("cells", {}).items():
            if str(value).startswith("="):
                formulas.append((value, refs))
            if len(formulas) >= 5:
                break
        if len(formulas) >= 5:
            break

    if formulas:
        print("\n  Sample formulas in encoding:")
        for val, refs in formulas:
            print(f"    {val}  -->  {refs}")


# =============================================================================
# Main
# =============================================================================


def run_test(name: str, creator_fn, path: str, k: int = 2) -> dict:
    """Create a test sheet, encode it, verify, and report."""
    print(f"\n{'='*60}")
    print(f"TEST: {name}")
    print(f"{'='*60}")

    # Create
    creator_fn(path)

    # Encode (timed)
    t0 = time.perf_counter()
    encoding = encode_spreadsheet(path, k=k)
    elapsed = time.perf_counter() - t0

    if not encoding:
        print(f"  [FAIL] Encoding returned None!")
        return {}

    print(f"  Encoding time: {elapsed:.3f}s")

    # Verify
    verify_structure(encoding, name)
    verify_formula_preservation(encoding, name)
    verify_ranges_merged(encoding, name)
    print_metrics(encoding, name)
    print_sample_formulas(encoding)

    # Also do vanilla for comparison
    t0 = time.perf_counter()
    vanilla = vanilla_encode(path)
    vanilla_time = time.perf_counter() - t0

    if vanilla:
        first_sheet = next(iter(vanilla))
        vanilla_tokens = len(vanilla[first_sheet])
        final_tokens = encoding.get("compression_metrics", {}).get("overall", {}).get("final_tokens", 0)
        print(f"\n  Vanilla tokens:   {vanilla_tokens:>10,}  (encoded in {vanilla_time:.3f}s)")
        if final_tokens > 0:
            print(f"  Compression vs vanilla: {vanilla_tokens / final_tokens:.2f}x")

        # Check formulas in vanilla too
        formula_in_vanilla = any("=" in line for line in vanilla[first_sheet].split("\n"))
        if formula_in_vanilla:
            print(f"  [PASS] Vanilla also preserves formulas")

    return encoding


def main() -> None:
    """Run all tests."""
    print("SpreadsheetLLM Encoder — Comprehensive Demo & Test")
    print("=" * 60)

    test_dir = os.path.join(os.path.dirname(__file__), "test_sheets")
    os.makedirs(test_dir, exist_ok=True)

    results = {}

    # Test 1: Tiny (5 rows)
    results["tiny"] = run_test(
        "Tiny (5 rows, basic formulas)",
        create_tiny_sheet,
        os.path.join(test_dir, "tiny.xlsx"),
    )

    # Test 2: Small (2 adjacent tables, no gap, merged header)
    results["small"] = run_test(
        "Small (2 adjacent tables, merged header, mixed types)",
        create_small_sheet,
        os.path.join(test_dir, "small.xlsx"),
    )

    # Test 3: Medium (500 rows, complex formulas)
    results["medium"] = run_test(
        "Medium (500 rows, 11 cols, complex formulas)",
        create_medium_sheet,
        os.path.join(test_dir, "medium.xlsx"),
    )

    # Test 4: Large (5000 rows, 3 tables, 3 sheets)
    results["large"] = run_test(
        "Large (5000 rows, 3 adjacent tables, 3 sheets)",
        create_large_sheet,
        os.path.join(test_dir, "large.xlsx"),
        k=2,
    )

    # --- Summary ---
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for name, enc in results.items():
        if enc:
            m = enc.get("compression_metrics", {}).get("overall", {})
            print(f"  {name:10s}  orig={m.get('original_tokens',0):>10,}  final={m.get('final_tokens',0):>8,}  ratio={m.get('overall_ratio',0):>6.2f}x")
        else:
            print(f"  {name:10s}  FAILED")

    print("\nDone.")


if __name__ == "__main__":
    main()
