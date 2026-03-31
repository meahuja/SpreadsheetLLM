"""Vanilla baseline encoding for spreadsheets.

Paper Section 3.1: "Vanilla serialization incorporating cell addresses,
values, and formats in a markdown-like style."

Format: row-major, pipe-delimited: A1,val|B1,val\nA2,val|B2,val
"""

from __future__ import annotations

import logging
import os
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def vanilla_encode(
    excel_path: str,
    output_path: Optional[str] = None,
) -> Optional[Dict[str, str]]:
    """Produce a simple vanilla markdown-like encoding of a spreadsheet.

    Args:
        excel_path: Path to .xlsx file.
        output_path: Optional path to save first sheet's encoding as text.

    Returns:
        Dict mapping sheet names to their vanilla encoding strings.
    """
    logger.info("Vanilla encoding: %s", excel_path)

    try:
        # data_only=False preserves formulas (preferred over computed values)
        workbook = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as exc:
        logger.error("Error loading workbook: %s", exc)
        return None

    result: Dict[str, str] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows_out = []

        for r in range(1, (sheet.max_row or 0) + 1):
            cells_out = []
            for c in range(1, (sheet.max_column or 0) + 1):
                cell = sheet.cell(row=r, column=c)
                ref = f"{get_column_letter(c)}{r}"
                val = str(cell.value) if cell.value is not None else ""
                cells_out.append(f"{ref},{val}")
            rows_out.append("|".join(cells_out))

        result[sheet_name] = "\n".join(rows_out)

    if output_path and result:
        first_sheet = next(iter(result))
        with open(output_path, "w", encoding="utf-8") as fh:
            fh.write(result[first_sheet])
        logger.info("Saved vanilla encoding to %s", output_path)

    return result
