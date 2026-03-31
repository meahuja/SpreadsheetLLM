"""SheetCompressor — 3-stage spreadsheet compression pipeline.

Implements the SheetCompressor framework from arXiv:2407.09025:
  Stage 1: Structural-Anchor-Based Extraction (Section 3.2)
  Stage 2: Inverted-Index Translation (Section 3.3)
  Stage 3: Data-Format-Aware Aggregation (Section 3.4)
"""

from __future__ import annotations

import json
import logging
import os
from collections import defaultdict
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string

from spreadsheet_llm.cell_utils import (
    cell_coord,
    detect_semantic_type,
    get_cell_style_fingerprint,
    get_number_format_string,
    infer_cell_data_type,
    split_cell_ref,
)

logger = logging.getLogger(__name__)

# --- Constants ---
MAX_CANDIDATES = 200
HEADER_THRESHOLD = 0.6
SPARSITY_THRESHOLD = 0.10
NMS_IOU_THRESHOLD = 0.5
HEADER_SCORE_WEIGHT = 10
MAX_BOUNDARY_ROWS = 100  # Cap boundary rows to prevent blowup


# =============================================================================
# Public API
# =============================================================================


def encode_spreadsheet(
    excel_path: str,
    k: int = 2,
    output_path: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """Encode an Excel file using the SheetCompressor pipeline.

    Args:
        excel_path: Path to .xlsx file.
        k: Neighborhood distance for anchor expansion (default 2).
        output_path: Optional path to write JSON output.

    Returns:
        Full encoding dict with sheets, compression_metrics.
    """
    logger.info("Processing: %s", excel_path)

    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=False)
    except FileNotFoundError:
        logger.error("File not found: %s", excel_path)
        return None
    except Exception as exc:
        logger.error("Error loading workbook: %s", exc)
        return None

    sheets_encoding: Dict[str, Any] = {}
    metrics: Dict[str, Any] = {"sheets": {}}
    totals = {"original": 0, "anchor": 0, "index": 0, "format": 0, "final": 0}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        if _is_empty_sheet(sheet):
            logger.info("Sheet '%s' is empty, skipping.", sheet_name)
            continue

        logger.info(
            "Sheet '%s': %d rows x %d cols",
            sheet_name, sheet.max_row, sheet.max_column,
        )

        # --- Original token count ---
        original_cells = _collect_all_cells(sheet)
        original_tokens = _token_count(original_cells)

        # --- Stage 1: Structural-Anchor-Based Extraction ---
        merged_map = _build_merged_cell_map(sheet)
        row_anchors, col_anchors = find_structural_anchors(sheet, k, merged_map)
        kept_rows, kept_cols = _expand_anchors(
            row_anchors, col_anchors, sheet.max_row, sheet.max_column, k,
        )
        kept_rows, kept_cols = compress_homogeneous_regions(
            sheet, kept_rows, kept_cols,
        )

        anchor_cells = _collect_kept_cells(sheet, kept_rows, kept_cols)
        anchor_tokens = _token_count(anchor_cells)

        # --- Stage 2: Inverted-Index Translation ---
        inverted_index, format_map = create_inverted_index(
            sheet, kept_rows, kept_cols, merged_map,
        )
        merged_index = create_inverted_index_translation(inverted_index)
        index_tokens = _token_count(merged_index)

        # --- Stage 3: Data-Format-Aware Aggregation ---
        type_nfs_groups = _group_by_semantic_type(sheet, format_map)
        aggregated_formats = aggregate_by_semantic_type(type_nfs_groups)
        format_tokens = _token_count(aggregated_formats)

        # Numeric sub-aggregation
        numeric_groups = {
            k_: v for k_, v in type_nfs_groups.items()
            if json.loads(k_).get("type") in ("integer", "float", "numeric")
        }
        numeric_ranges = aggregate_by_semantic_type(numeric_groups)

        # --- Assemble sheet encoding ---
        sheet_encoding = {
            "structural_anchors": {
                "rows": sorted(row_anchors),
                "columns": [get_column_letter(c) for c in sorted(col_anchors)],
            },
            "cells": merged_index,
            "formats": aggregated_formats,
            "numeric_ranges": numeric_ranges,
        }
        final_tokens = _token_count(sheet_encoding)

        sheets_encoding[sheet_name] = sheet_encoding
        sheet_metrics = _compute_metrics(
            original_tokens, anchor_tokens, index_tokens, format_tokens, final_tokens,
        )
        metrics["sheets"][sheet_name] = sheet_metrics

        for key, val in zip(
            ("original", "anchor", "index", "format", "final"),
            (original_tokens, anchor_tokens, index_tokens, format_tokens, final_tokens),
        ):
            totals[key] += val

        logger.info(
            "%s — Anchor: %.1fx, Index: %.1fx, Format: %.1fx, Overall: %.1fx",
            sheet_name,
            sheet_metrics["anchor_ratio"],
            sheet_metrics["inverted_index_ratio"],
            sheet_metrics["format_ratio"],
            sheet_metrics["overall_ratio"],
        )

    metrics["overall"] = _compute_metrics(
        totals["original"], totals["anchor"], totals["index"],
        totals["format"], totals["final"],
    )

    result = {
        "file_name": os.path.basename(excel_path),
        "sheets": sheets_encoding,
        "compression_metrics": metrics,
    }

    if output_path:
        with open(output_path, "w", encoding="utf-8") as fh:
            json.dump(result, fh, indent=2, ensure_ascii=False)
        logger.info("Saved encoding to %s", output_path)

    return result


# =============================================================================
# Stage 1: Structural-Anchor-Based Extraction
# =============================================================================


def _build_merged_cell_map(sheet: Any) -> Dict[str, str]:
    """Build {cell_coordinate: merge_start_coordinate} for O(1) lookups."""
    merged_map: Dict[str, str] = {}
    for merge_range in sheet.merged_cells.ranges:
        start = merge_range.start_cell.coordinate
        for row, col in merge_range.cells:
            coord = cell_coord(row, col)
            merged_map[coord] = start
    return merged_map


def _analyze_rows_single_pass(
    sheet: Any,
    merged_map: Dict[str, str],
) -> Dict[str, Any]:
    """Single-pass analysis of all rows: fingerprints, emptiness, widths,
    header detection, data type composition.

    Returns a dict with all collected info, avoiding multiple full-sheet scans.
    """
    max_row = sheet.max_row
    max_col = sheet.max_column

    row_fingerprints: List[int] = []      # hash of row profile
    row_empty: List[bool] = []            # is row fully empty?
    row_widths: List[Tuple[int, int]] = []  # (min_col, max_col) of populated
    row_is_header: List[bool] = []        # header heuristic result
    row_numeric_count: List[int] = []     # count of numeric cells
    row_text_count: List[int] = []        # count of text cells
    row_populated_count: List[int] = []   # count of non-empty cells

    for r in range(1, max_row + 1):
        fp_parts = []
        is_empty = True
        min_c = max_col + 1
        max_c = 0
        populated = 0
        bold_count = 0
        center_count = 0
        caps_count = 0
        border_count = 0
        string_count = 0
        numeric = 0
        text = 0

        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            coord = cell_coord(r, c)
            is_merged = coord in merged_map
            val = cell.value
            val_hash = hash(val) if val is not None else 0
            style_id = get_cell_style_fingerprint(cell)
            fp_parts.append((val_hash, is_merged, style_id))

            if val is not None and str(val).strip() != "":
                is_empty = False
                populated += 1
                min_c = min(min_c, c)
                max_c = max(max_c, c)

                # Header heuristics
                if cell.font and cell.font.bold:
                    bold_count += 1
                if cell.alignment and cell.alignment.horizontal == "center":
                    center_count += 1
                if (cell.border and cell.border.bottom
                        and cell.border.bottom.style
                        and cell.border.bottom.style != "none"):
                    border_count += 1
                if isinstance(val, str):
                    string_count += 1
                    if val.isupper() and len(val) > 1:
                        caps_count += 1

                # Data type
                dt = cell.data_type
                if dt == "n":
                    numeric += 1
                elif dt == "s":
                    text += 1

        row_fingerprints.append(hash(tuple(fp_parts)))
        row_empty.append(is_empty)
        row_widths.append((min_c, max_c))
        row_populated_count.append(populated)
        row_numeric_count.append(numeric)
        row_text_count.append(text)

        # Header detection
        is_header = False
        if populated > 0:
            if bold_count / populated > HEADER_THRESHOLD:
                is_header = True
            elif center_count / populated > HEADER_THRESHOLD:
                is_header = True
            elif border_count / populated > HEADER_THRESHOLD:
                is_header = True
            elif string_count > 0 and caps_count / string_count > HEADER_THRESHOLD:
                is_header = True
        row_is_header.append(is_header)

    return {
        "fingerprints": row_fingerprints,
        "empty": row_empty,
        "widths": row_widths,
        "is_header": row_is_header,
        "numeric_count": row_numeric_count,
        "text_count": row_text_count,
        "populated_count": row_populated_count,
    }


def _analyze_cols_single_pass(
    sheet: Any,
    merged_map: Dict[str, str],
) -> List[int]:
    """Build column fingerprint hashes in a single pass."""
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Build column data transposed
    col_parts: List[List[Any]] = [[] for _ in range(max_col)]

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            coord = cell_coord(r, c)
            is_merged = coord in merged_map
            val = cell.value
            val_hash = hash(val) if val is not None else 0
            style_id = get_cell_style_fingerprint(cell)
            col_parts[c - 1].append((val_hash, is_merged, style_id))

    return [hash(tuple(parts)) for parts in col_parts]


def find_boundary_candidates(
    sheet: Any,
    merged_map: Dict[str, str],
) -> Tuple[List[int], List[int]]:
    """Find row/column boundary candidates via single-pass analysis.

    Returns (row_boundaries, col_boundaries) — raw boundary indices.
    """
    row_info = _analyze_rows_single_pass(sheet, merged_map)
    col_fps = _analyze_cols_single_pass(sheet, merged_map)

    fps = row_info["fingerprints"]
    empty = row_info["empty"]
    widths = row_info["widths"]
    is_header = row_info["is_header"]
    num_counts = row_info["numeric_count"]
    text_counts = row_info["text_count"]

    max_row = sheet.max_row

    row_boundaries: Set[int] = set()

    for r_idx in range(len(fps)):
        r = r_idx + 1  # 1-indexed

        # Fingerprint difference with previous row
        if r_idx > 0 and fps[r_idx] != fps[r_idx - 1]:
            row_boundaries.add(r)
            row_boundaries.add(r - 1 if r > 1 else r)

        # Empty row → boundary on both sides
        if empty[r_idx]:
            row_boundaries.add(r)
            if r > 1:
                row_boundaries.add(r - 1)
            if r < max_row:
                row_boundaries.add(r + 1)

        # Column width transition
        if r_idx > 0 and widths[r_idx] != widths[r_idx - 1]:
            row_boundaries.add(r)

        # Header row → boundary
        if is_header[r_idx]:
            row_boundaries.add(r)
            if r > 1:
                row_boundaries.add(r - 1)

        # Data type transition (numeric-dominant ↔ text-dominant)
        if r_idx > 0:
            prev_n, prev_t = num_counts[r_idx - 1], text_counts[r_idx - 1]
            curr_n, curr_t = num_counts[r_idx], text_counts[r_idx]
            if (prev_n + prev_t) >= 2 and (curr_n + curr_t) >= 2:
                prev_dom = "n" if prev_n > prev_t else "t"
                curr_dom = "n" if curr_n > curr_t else "t"
                if prev_dom != curr_dom:
                    row_boundaries.add(r)

    # Column boundaries from fingerprint differences
    col_boundaries: Set[int] = set()
    for c_idx in range(1, len(col_fps)):
        if col_fps[c_idx] != col_fps[c_idx - 1]:
            col_boundaries.add(c_idx + 1)  # 1-indexed
            col_boundaries.add(c_idx)

    # Clamp
    row_boundaries = {r for r in row_boundaries if 1 <= r <= max_row}
    col_boundaries = {c for c in col_boundaries if 1 <= c <= sheet.max_column}

    # Cap row boundaries to prevent combinatorial explosion
    sorted_rows = sorted(row_boundaries)
    if len(sorted_rows) > MAX_BOUNDARY_ROWS:
        # Keep evenly spaced subset + first/last + all header rows
        header_rows = {r for r in sorted_rows if is_header[r - 1]}
        step = len(sorted_rows) // MAX_BOUNDARY_ROWS
        sampled = set(sorted_rows[::step])
        sampled.add(sorted_rows[0])
        sampled.add(sorted_rows[-1])
        sampled |= header_rows
        sorted_rows = sorted(sampled)

    return sorted_rows, sorted(col_boundaries)


def _compose_candidates_consecutive(
    row_bounds: List[int],
    col_bounds: List[int],
) -> List[Tuple[int, int, int, int]]:
    """Compose rectangular candidates using CONSECUTIVE boundary pairs only.

    This is O(row_bounds × col_bounds) instead of O(n^4).
    For each consecutive row pair and consecutive col pair, create a candidate.
    Also create a few spanning candidates (first row to each subsequent).
    """
    candidates = []

    if len(row_bounds) < 2 or len(col_bounds) < 2:
        # Single boundary — use sheet extents
        if row_bounds and col_bounds:
            candidates.append((
                row_bounds[0], col_bounds[0],
                row_bounds[-1], col_bounds[-1],
            ))
        return candidates

    # Consecutive row pairs × consecutive col pairs
    for i in range(len(row_bounds) - 1):
        r1, r2 = row_bounds[i], row_bounds[i + 1]
        for j in range(len(col_bounds) - 1):
            c1, c2 = col_bounds[j], col_bounds[j + 1]
            candidates.append((r1, c1, r2, c2))

    # Also try spanning from first row boundary to each subsequent
    # (catches tables that span multiple boundary segments)
    r_first = row_bounds[0]
    c_first = col_bounds[0]
    c_last = col_bounds[-1]
    for i in range(2, min(len(row_bounds), 10)):
        candidates.append((r_first, c_first, row_bounds[i], c_last))

    # And full-span candidate
    candidates.append((row_bounds[0], col_bounds[0], row_bounds[-1], col_bounds[-1]))

    # Deduplicate
    candidates = list(set(candidates))

    if len(candidates) > MAX_CANDIDATES:
        candidates.sort(
            key=lambda x: (x[2] - x[0]) * (x[3] - x[1]),
            reverse=True,
        )
        candidates = candidates[:MAX_CANDIDATES]

    return candidates


def _filter_candidates(
    sheet: Any,
    candidates: List[Tuple[int, int, int, int]],
    row_info: Dict[str, Any],
) -> List[Tuple[int, int, int, int]]:
    """Filter candidates by size, sparsity, and header presence.

    Uses pre-computed row_info to avoid re-scanning cells.
    """
    populated_counts = row_info["populated_count"]
    is_header = row_info["is_header"]
    filtered = []

    for r1, c1, r2, c2 in candidates:
        if r2 - r1 < 1 or c2 - c1 < 1:
            continue

        # Estimate populated count from row-level data (approximate but fast)
        total_cells = (r2 - r1 + 1) * (c2 - c1 + 1)
        est_populated = sum(
            populated_counts[r - 1] for r in range(r1, r2 + 1)
            if r - 1 < len(populated_counts)
        )
        # Scale by column fraction (row populated count is for full width)
        col_fraction = (c2 - c1 + 1) / sheet.max_column if sheet.max_column > 0 else 1
        est_populated = int(est_populated * col_fraction)

        if total_cells > 0 and est_populated / total_cells < SPARSITY_THRESHOLD:
            continue

        # Header presence (from pre-computed data)
        has_header = any(
            is_header[r - 1]
            for r in range(r1, min(r1 + 3, r2 + 1))
            if r - 1 < len(is_header)
        )

        if not has_header and total_cells > 0:
            # Fallback: accept if >50% populated and text-dominant first row
            text_counts = row_info["text_count"]
            num_counts = row_info["numeric_count"]
            if est_populated / total_cells > 0.5:
                if (r1 - 1 < len(text_counts)
                        and text_counts[r1 - 1] > num_counts[r1 - 1]):
                    has_header = True

        if not has_header:
            continue

        filtered.append((r1, c1, r2, c2))

    return filtered


def _calculate_iou(
    box1: Tuple[int, int, int, int],
    box2: Tuple[int, int, int, int],
) -> float:
    """Calculate Intersection over Union for two bounding boxes."""
    r1_1, c1_1, r2_1, c2_1 = box1
    r1_2, c1_2, r2_2, c2_2 = box2

    inter_r1 = max(r1_1, r1_2)
    inter_c1 = max(c1_1, c1_2)
    inter_r2 = min(r2_1, r2_2)
    inter_c2 = min(c2_1, c2_2)

    inter_area = max(0, inter_r2 - inter_r1 + 1) * max(0, inter_c2 - inter_c1 + 1)
    area1 = (r2_1 - r1_1 + 1) * (c2_1 - c1_1 + 1)
    area2 = (r2_2 - r1_2 + 1) * (c2_2 - c1_2 + 1)
    union_area = area1 + area2 - inter_area

    return inter_area / union_area if union_area > 0 else 0.0


def _nms_candidates(
    candidates: List[Tuple[int, int, int, int]],
    row_info: Dict[str, Any],
) -> List[Tuple[int, int, int, int]]:
    """Non-Maximum Suppression using pre-computed row info for scoring."""
    if not candidates:
        return []

    is_header = row_info["is_header"]
    populated_counts = row_info["populated_count"]

    scores = []
    for r1, c1, r2, c2 in candidates:
        score = 0
        for r in range(r1, min(r1 + 3, r2 + 1)):
            if r - 1 < len(is_header) and is_header[r - 1]:
                score += HEADER_SCORE_WEIGHT
        # Use pre-computed populated counts (approximate area score)
        for r in range(r1, r2 + 1):
            if r - 1 < len(populated_counts):
                score += populated_counts[r - 1]
        scores.append(score)

    indices = sorted(range(len(candidates)), key=lambda i: scores[i], reverse=True)

    keep = []
    while indices:
        best = indices.pop(0)
        keep.append(best)
        remaining = []
        for idx in indices:
            if _calculate_iou(candidates[best], candidates[idx]) < NMS_IOU_THRESHOLD:
                remaining.append(idx)
        indices = remaining

    return [candidates[i] for i in keep]


def find_structural_anchors(
    sheet: Any,
    k: int,
    merged_map: Dict[str, str],
) -> Tuple[List[int], List[int]]:
    """Find structural anchors: single-pass analysis → boundary detection →
    consecutive-pair candidates → filtering → NMS → anchor extraction.

    Falls back to keeping all rows/cols if no candidates survive.
    """
    # Single-pass row analysis (collects fingerprints, emptiness, widths,
    # headers, data types all at once)
    row_info = _analyze_rows_single_pass(sheet, merged_map)

    row_bounds, col_bounds = find_boundary_candidates(sheet, merged_map)

    if not row_bounds or not col_bounds:
        return (
            list(range(1, sheet.max_row + 1)),
            list(range(1, sheet.max_column + 1)),
        )

    candidates = _compose_candidates_consecutive(row_bounds, col_bounds)
    candidates = _filter_candidates(sheet, candidates, row_info)
    candidates = _nms_candidates(candidates, row_info)

    if not candidates:
        return (
            list(range(1, sheet.max_row + 1)),
            list(range(1, sheet.max_column + 1)),
        )

    # Extract anchors
    row_anchors: Set[int] = set()
    col_anchors: Set[int] = set()
    for r1, c1, r2, c2 in candidates:
        row_anchors.add(r1)
        row_anchors.add(r2)
        col_anchors.add(c1)
        col_anchors.add(c2)

    # Add header rows from pre-computed data
    for r_idx, is_h in enumerate(row_info["is_header"]):
        if is_h:
            row_anchors.add(r_idx + 1)

    return sorted(row_anchors), sorted(col_anchors)


def _expand_anchors(
    row_anchors: List[int],
    col_anchors: List[int],
    max_row: int,
    max_col: int,
    k: int,
) -> Tuple[List[int], List[int]]:
    """Expand anchors by k-neighborhood."""
    kept_rows: Set[int] = set()
    for anchor in row_anchors:
        for i in range(max(1, anchor - k), min(max_row + 1, anchor + k + 1)):
            kept_rows.add(i)

    kept_cols: Set[int] = set()
    for anchor in col_anchors:
        for i in range(max(1, anchor - k), min(max_col + 1, anchor + k + 1)):
            kept_cols.add(i)

    return sorted(kept_rows), sorted(kept_cols)


def compress_homogeneous_regions(
    sheet: Any,
    rows: List[int],
    cols: List[int],
) -> Tuple[List[int], List[int]]:
    """Remove rows/cols where all kept cells are identical (value + format).

    Uses early-exit for performance.
    """
    def _row_is_homogeneous(r: int) -> bool:
        first_val = None
        first_fmt = None
        for c in cols:
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            fmt = cell.number_format
            if first_val is None and first_fmt is None:
                first_val = val
                first_fmt = fmt
            elif val != first_val or fmt != first_fmt:
                return False
        return True

    def _col_is_homogeneous(c: int) -> bool:
        first_val = None
        first_fmt = None
        for r in rows:
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            fmt = cell.number_format
            if first_val is None and first_fmt is None:
                first_val = val
                first_fmt = fmt
            elif val != first_val or fmt != first_fmt:
                return False
        return True

    kept_rows = [r for r in rows if not _row_is_homogeneous(r)]
    kept_cols = [c for c in cols if not _col_is_homogeneous(c)]

    if not kept_rows:
        kept_rows = rows
    if not kept_cols:
        kept_cols = cols

    return kept_rows, kept_cols


# =============================================================================
# Stage 2: Inverted-Index Translation
# =============================================================================


def create_inverted_index(
    sheet: Any,
    rows: List[int],
    cols: List[int],
    merged_map: Dict[str, str],
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    """Create inverted index: value → [cell_refs] and format_key → [cell_refs].

    Handles merged cells by redirecting to merge start cell.
    Skips empty cells. Preserves formulas (data_only=False).
    """
    inverted: Dict[str, List[str]] = defaultdict(list)
    format_groups: Dict[str, List[str]] = defaultdict(list)

    for r in rows:
        for c in cols:
            coord = cell_coord(r, c)
            cell = sheet.cell(row=r, column=c)

            # Resolve merged cell value
            if coord in merged_map:
                start_coord = merged_map[coord]
                value = sheet[start_coord].value
            else:
                value = cell.value

            if value is None:
                continue

            val_str = str(value)
            inverted[val_str].append(coord)

            sem_type = detect_semantic_type(cell)
            nfs = get_number_format_string(cell)
            fmt_key = json.dumps({"type": sem_type, "nfs": nfs}, sort_keys=True)
            format_groups[fmt_key].append(coord)

    return dict(inverted), dict(format_groups)


def merge_cell_ranges(refs: List[str]) -> List[str]:
    """Merge a list of cell references into maximal rectangular ranges.

    Greedy algorithm: for each unprocessed cell, expand right then down.
    Uses set membership for O(1) lookups.
    """
    if not refs:
        return []

    coords = set()
    for ref in refs:
        try:
            col_letter, row = split_cell_ref(ref)
            col = column_index_from_string(col_letter)
            coords.add((row, col))
        except Exception:
            continue

    if not coords:
        return []

    processed: Set[Tuple[int, int]] = set()
    ranges: List[str] = []

    for row, col in sorted(coords):
        if (row, col) in processed:
            continue

        # Expand right
        width = 1
        while (row, col + width) in coords and (row, col + width) not in processed:
            width += 1

        # Expand down (full-width rows only)
        height = 1
        expanding = True
        while expanding:
            next_row = row + height
            for w in range(width):
                if (next_row, col + w) not in coords or (next_row, col + w) in processed:
                    expanding = False
                    break
            if expanding:
                height += 1

        end_row = row + height - 1
        end_col = col + width - 1
        start_ref = cell_coord(row, col)
        end_ref = cell_coord(end_row, end_col)

        if width == 1 and height == 1:
            ranges.append(start_ref)
        else:
            ranges.append(f"{start_ref}:{end_ref}")

        for r in range(row, row + height):
            for c in range(col, col + width):
                processed.add((r, c))

    return ranges


def create_inverted_index_translation(
    inverted_index: Dict[str, List[str]],
) -> Dict[str, List[str]]:
    """Apply range merging to each value's cell references."""
    return {
        value: merge_cell_ranges(refs)
        for value, refs in inverted_index.items()
        if value and str(value).strip()
    }


# =============================================================================
# Stage 3: Data-Format-Aware Aggregation
# =============================================================================


def _group_by_semantic_type(
    sheet: Any,
    format_map: Dict[str, List[str]],
) -> Dict[str, List[str]]:
    """Re-group format_map entries by (semantic_type, NFS) key."""
    groups: Dict[str, List[str]] = defaultdict(list)
    for _fmt_key, refs in format_map.items():
        for ref in refs:
            try:
                cell = sheet[ref]
            except Exception:
                continue
            sem_type = detect_semantic_type(cell)
            nfs = get_number_format_string(cell)
            key = json.dumps({"type": sem_type, "nfs": nfs}, sort_keys=True)
            groups[key].append(ref)
    return dict(groups)


def aggregate_by_semantic_type(
    type_nfs_groups: Dict[str, List[str]],
) -> Dict[str, List[str]]:
    """Find contiguous rectangular regions for each semantic type group."""
    return {
        key: merge_cell_ranges(refs)
        for key, refs in type_nfs_groups.items()
        if refs
    }


# =============================================================================
# Helpers
# =============================================================================


def _is_empty_sheet(sheet: Any) -> bool:
    """Check if a sheet is effectively empty."""
    if sheet.max_row is None or sheet.max_column is None:
        return True
    if sheet.max_row <= 1 and sheet.max_column <= 1:
        if sheet.cell(row=1, column=1).value is None:
            return True
    return False


def _collect_all_cells(sheet: Any) -> Dict[str, str]:
    """Collect all non-empty cells for token counting."""
    cells: Dict[str, str] = {}
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                cells[cell_coord(r, c)] = str(val)
    return cells


def _collect_kept_cells(
    sheet: Any,
    rows: List[int],
    cols: List[int],
) -> Dict[str, str]:
    """Collect non-empty cells in kept rows/cols."""
    cells: Dict[str, str] = {}
    for r in rows:
        for c in cols:
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                cells[cell_coord(r, c)] = str(val)
    return cells


def _token_count(data: Any) -> int:
    """Proxy for token count: length of JSON serialization."""
    return len(json.dumps(data, ensure_ascii=False))


def _compression_ratio(original: int, compressed: int) -> float:
    """Calculate compression ratio (higher = more compression)."""
    if compressed == 0:
        return 0.0
    return original / compressed


def _compute_metrics(
    original: int,
    anchor: int,
    index: int,
    fmt: int,
    final: int,
) -> Dict[str, Any]:
    """Compute compression metrics for all stages."""
    return {
        "original_tokens": original,
        "after_anchor_tokens": anchor,
        "after_inverted_index_tokens": index,
        "after_format_tokens": fmt,
        "final_tokens": final,
        "anchor_ratio": _compression_ratio(original, anchor),
        "inverted_index_ratio": _compression_ratio(original, index),
        "format_ratio": _compression_ratio(original, fmt),
        "overall_ratio": _compression_ratio(original, final),
    }
