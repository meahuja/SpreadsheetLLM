"""Create a messy 2000-row Excel with 4 tables, encode it, print JSON."""
import json
import os
import sys
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from spreadsheet_llm.encoder import encode_spreadsheet

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")

path = os.path.join(os.path.dirname(__file__), "messy_2000.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Messy"

# ==========================================================================
# TABLE 1: Financial transactions (rows 1-501)
#   - Merged title row, bold headers, currency/pct/date formats
#   - Some cells have errors (#N/A text), some empty columns scattered
#   - Formulas: SUBTOTAL, running balance
# ==========================================================================
ws.merge_cells("A1:J1")
ws.cell(row=1, column=1, value="Q1-Q4 TRANSACTION LEDGER 2024").font = Font(bold=True, size=16, color="FFFFFF")
ws.cell(row=1, column=1).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ws.cell(row=1, column=1).alignment = CENTER

headers1 = ["TXN_ID", "Date", "", "Vendor", "Category", "Amount", "Tax%", "Tax Amt", "Total", "Running Bal"]
for c, h in enumerate(headers1, 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    cell.border = Border(bottom=Side(style="medium"))

vendors = ["Acme Corp", "Globex Inc", "Initech LLC", "Umbrella Co", "Stark Ind",
           "Wayne Ent", "Oscorp", "LexCorp", "Capsule Corp", "Cyberdyne"]
categories = ["Office Supplies", "IT Equipment", "Travel", "Marketing", "Legal",
              "Consulting", "SaaS", "Hardware", "Rent", "Utilities"]

for r in range(3, 503):
    idx = r - 3
    ws.cell(row=r, column=1, value=f"TXN-{idx+1:05d}")
    ws.cell(row=r, column=2, value=date(2024, (idx % 12) + 1, (idx % 28) + 1)).number_format = "dd-mmm-yyyy"
    # Column C intentionally blank (messy — empty column in middle)
    ws.cell(row=r, column=4, value=vendors[idx % 10])

    # Scatter some #N/A errors and weird values
    if idx % 47 == 0:
        ws.cell(row=r, column=5, value="#N/A")
    elif idx % 31 == 0:
        ws.cell(row=r, column=5, value="UNCATEGORIZED")
    else:
        ws.cell(row=r, column=5, value=categories[idx % 10])

    amt = round(50 + (idx % 200) * 12.75 + (idx % 7) * 0.99, 2)
    ws.cell(row=r, column=6, value=amt).number_format = '"$"#,##0.00'

    tax_rate = 0.08 if idx % 3 == 0 else (0.10 if idx % 3 == 1 else 0.065)
    ws.cell(row=r, column=7, value=tax_rate).number_format = "0.00%"
    ws.cell(row=r, column=8, value=f"=F{r}*G{r}").number_format = '"$"#,##0.00'
    ws.cell(row=r, column=9, value=f"=F{r}+H{r}").number_format = '"$"#,##0.00'

    # Running balance formula
    if r == 3:
        ws.cell(row=r, column=10, value=f"=I{r}").number_format = '"$"#,##0.00'
    else:
        ws.cell(row=r, column=10, value=f"=J{r-1}+I{r}").number_format = '"$"#,##0.00'

# Subtotals row
r = 503
ws.cell(row=r, column=1, value="SUBTOTAL").font = Font(bold=True, size=11)
ws.cell(row=r, column=6, value="=SUBTOTAL(9,F3:F502)").number_format = '"$"#,##0.00'
ws.cell(row=r, column=8, value="=SUBTOTAL(9,H3:H502)").number_format = '"$"#,##0.00'
ws.cell(row=r, column=9, value="=SUBTOTAL(9,I3:I502)").number_format = '"$"#,##0.00'
# Blank row 504

# ==========================================================================
# TABLE 2: Employee roster (rows 505-855) — NO GAP from table 1
#   - Different column width, different header color
#   - Mix of dates, emails, booleans, text
#   - Some rows have missing data (sparse)
# ==========================================================================
headers2 = ["EMP_ID", "First", "Last", "Email", "Dept", "Hire Date", "Salary",
            "Active", "Manager", "Notes"]
for c, h in enumerate(headers2, 1):
    cell = ws.cell(row=505, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = CENTER

firsts = ["Alice", "Bob", "Carol", "Dave", "Eve", "Frank", "Grace", "Hank", "Ivy", "Jack"]
lasts = ["Smith", "Jones", "White", "Brown", "Black", "Green", "Adams", "Clark", "Davis", "Evans"]
depts = ["Engineering", "Sales", "Marketing", "HR", "Finance", "Legal", "Ops", "Support"]

for r in range(506, 856):
    idx = r - 506
    ws.cell(row=r, column=1, value=f"EMP-{idx+1:04d}")
    ws.cell(row=r, column=2, value=firsts[idx % 10])
    ws.cell(row=r, column=3, value=lasts[idx % 10])

    email = f"{firsts[idx%10].lower()}.{lasts[idx%10].lower()}@company.com"
    ws.cell(row=r, column=4, value=email)

    # Sparse: skip dept for every 13th row
    if idx % 13 != 0:
        ws.cell(row=r, column=5, value=depts[idx % 8])

    ws.cell(row=r, column=6, value=date(2015 + (idx % 10), (idx % 12) + 1, 1)).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=7, value=55000 + (idx % 50) * 1500).number_format = '"$"#,##0'

    # Boolean active flag — some False
    ws.cell(row=r, column=8, value=idx % 17 != 0)

    # Manager reference — cross-row formula
    if idx > 0 and idx % 5 == 0:
        ws.cell(row=r, column=9, value=f"=B{506 + (idx // 5) - 1}&\" \"&C{506 + (idx // 5) - 1}")
    else:
        ws.cell(row=r, column=9, value="N/A")

    # Random notes — sparse
    if idx % 7 == 0:
        ws.cell(row=r, column=10, value=f"Performance review due {2025 + idx % 3}")

# ==========================================================================
# TABLE 3: Inventory (rows 856-1456) — 1 blank row gap
#   - Scientific notation, mixed units, nested IF formulas
#   - Some merged cells for category headers
# ==========================================================================
# Row 856 blank (gap)

ws.merge_cells("A857:H857")
ws.cell(row=857, column=1, value="WAREHOUSE INVENTORY - LIVE STOCK LEVELS").font = Font(bold=True, size=14)
ws.cell(row=857, column=1).fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

headers3 = ["SKU", "Product", "Category", "Qty", "Unit Price", "Value", "Reorder Pt", "Status"]
for c, h in enumerate(headers3, 1):
    cell = ws.cell(row=858, column=c, value=h)
    cell.font = BOLD
    cell.border = Border(bottom=Side(style="double"))

products = ["Widget-A", "Widget-B", "Gadget-X", "Gadget-Y", "Sensor-Z",
            "Module-P", "Board-Q", "Cable-R", "Adapter-S", "Chip-T"]
inv_categories = ["Electronics", "Mechanical", "Sensors", "Cables", "ICs"]

for r in range(859, 1459):
    idx = r - 859
    ws.cell(row=r, column=1, value=f"SKU-{idx+10000}")
    ws.cell(row=r, column=2, value=products[idx % 10])
    ws.cell(row=r, column=3, value=inv_categories[idx % 5])

    qty = (idx % 500) + 1
    ws.cell(row=r, column=4, value=qty).number_format = "#,##0"

    price = round(0.05 + (idx % 100) * 0.47, 2)
    ws.cell(row=r, column=5, value=price).number_format = '"$"#,##0.00'
    ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = '"$"#,##0.00'

    reorder = 50 + (idx % 100)
    ws.cell(row=r, column=7, value=reorder)

    # Nested IF formula for status
    ws.cell(row=r, column=8, value=f'=IF(D{r}=0,"OUT OF STOCK",IF(D{r}<G{r},"LOW","OK"))')

# Inventory summary
r = 1459
ws.cell(row=r, column=1, value="TOTAL SKUs").font = BOLD
ws.cell(row=r, column=4, value="=SUM(D859:D1458)").number_format = "#,##0"
ws.cell(row=r, column=6, value="=SUM(F859:F1458)").number_format = '"$"#,##0.00'

r = 1460
ws.cell(row=r, column=1, value="AVG Price")
ws.cell(row=r, column=5, value="=AVERAGE(E859:E1458)").number_format = '"$"#,##0.00'
ws.cell(row=r, column=1).font = BOLD

r = 1461
ws.cell(row=r, column=1, value="Low Stock Items")
ws.cell(row=r, column=4, value='=COUNTIF(H859:H1458,"LOW")').number_format = "#,##0"
ws.cell(row=r, column=1).font = BOLD

# ==========================================================================
# TABLE 4: Metrics dashboard (rows 1463-2000) — NO GAP, same columns as table 3
#   - Time series data, scientific notation values
#   - VLOOKUP, INDEX/MATCH, complex nested formulas
#   - Percentage changes, conditional formatting indicators
# ==========================================================================
headers4 = ["Timestamp", "Sensor", "Raw Value", "Calibrated", "Unit",
            "Delta%", "Alarm", "Site"]
for c, h in enumerate(headers4, 1):
    cell = ws.cell(row=1462, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF", italic=True)
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = CENTER

sensors = ["TEMP-01", "TEMP-02", "HUM-01", "PRESS-01", "FLOW-01",
           "PH-01", "TURB-01", "DO-01"]
units = ["°C", "°C", "%RH", "kPa", "L/min", "pH", "NTU", "mg/L"]
sites = ["Plant-A", "Plant-B", "Lab-1", "Lab-2", "Field-Station"]

for r in range(1463, 2001):
    idx = r - 1463
    hour = idx % 24
    minute = (idx * 5) % 60
    ws.cell(row=r, column=1, value=datetime(2024, 6, 1 + (idx // 24) % 28, hour, minute)).number_format = "yyyy-mm-dd hh:mm:ss"

    sensor_idx = idx % 8
    ws.cell(row=r, column=2, value=sensors[sensor_idx])

    # Raw value — some in scientific notation range
    if sensor_idx in (3, 6):  # pressure, turbidity — large/small numbers
        raw = (101.325 + (idx % 50) * 0.1) * 1000 if sensor_idx == 3 else 0.001 * (idx % 100)
        ws.cell(row=r, column=3, value=raw).number_format = "0.00E+00"
    else:
        raw = round(20.0 + (idx % 80) * 0.25, 2)
        ws.cell(row=r, column=3, value=raw)

    # Calibrated = raw * calibration factor (formula)
    ws.cell(row=r, column=4, value=f"=C{r}*1.002+0.05")

    ws.cell(row=r, column=5, value=units[sensor_idx])

    # Delta% from previous reading of same sensor
    if r > 1470:
        ws.cell(row=r, column=6, value=f"=IF(C{r-8}=0,0,(C{r}-C{r-8})/C{r-8})").number_format = "0.00%"
    else:
        ws.cell(row=r, column=6, value=0).number_format = "0.00%"

    # Alarm: nested IF on thresholds
    ws.cell(row=r, column=7, value=f'=IF(ABS(F{r})>0.1,"CRITICAL",IF(ABS(F{r})>0.05,"WARNING","OK"))')

    ws.cell(row=r, column=8, value=sites[idx % 5])

# Final stats rows
r = 2001
ws.cell(row=r, column=1, value="SENSOR STATS").font = Font(bold=True, size=12)
ws.merge_cells(f"A{r}:H{r}")

for i, sensor in enumerate(sensors):
    r = 2002 + i
    ws.cell(row=r, column=1, value=sensor).font = BOLD
    ws.cell(row=r, column=2, value="Readings")
    ws.cell(row=r, column=3, value=f'=COUNTIF(B1463:B2000,"{sensor}")')
    ws.cell(row=r, column=4, value="Avg")
    ws.cell(row=r, column=5, value=f'=AVERAGEIF(B1463:B2000,"{sensor}",C1463:C2000)')
    ws.cell(row=r, column=6, value="Alarms")
    ws.cell(row=r, column=7, value=f'=COUNTIFS(B1463:B2000,"{sensor}",G1463:G2000,"CRITICAL")')

wb.save(path)
print(f"Created: {path}")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

# ==========================================================================
# ENCODE IT
# ==========================================================================
print("\nEncoding...")
import time
t0 = time.perf_counter()
result = encode_spreadsheet(path, k=2, output_path=path.replace(".xlsx", "_encoded.json"))
elapsed = time.perf_counter() - t0
print(f"Encoding time: {elapsed:.2f}s")

# Print full JSON
print("\n" + "="*70)
print("FULL COMPRESSED JSON OUTPUT")
print("="*70)
print(json.dumps(result, indent=2, ensure_ascii=False))
