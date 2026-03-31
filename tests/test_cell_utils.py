"""Unit tests for cell_utils.py — semantic type detection and helpers."""
from __future__ import annotations

import os
import pytest
import openpyxl
from datetime import date, datetime

from spreadsheet_llm.cell_utils import (
    infer_cell_data_type,
    get_number_format_string,
    categorize_number_format,
    detect_semantic_type,
    get_cell_style_fingerprint,
    cell_coord,
    split_cell_ref,
)


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a temp workbook, yield (workbook, sheet, path), save on exit."""
    path = str(tmp_path / "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    yield wb, ws, path
    wb.save(path)


class TestInferCellDataType:
    def test_empty_cell(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "empty"

    def test_text_cell(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="hello")
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "text"

    def test_numeric_cell(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value=42)
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "numeric"

    def test_float_cell(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value=3.14)
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "numeric"

    def test_boolean_cell(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value=True)
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "boolean"

    def test_email_cell(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="test@example.com")
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "email"

    def test_non_email_text(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="not-an-email")
        assert infer_cell_data_type(ws.cell(row=1, column=1)) == "text"


class TestDetectSemanticType:
    def test_integer(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value=42)
        assert detect_semantic_type(ws.cell(row=1, column=1)) == "integer"

    def test_float(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value=3.14)
        assert detect_semantic_type(ws.cell(row=1, column=1)) == "float"

    def test_percentage(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        cell = ws.cell(row=1, column=1, value=0.85)
        cell.number_format = "0.00%"
        assert detect_semantic_type(cell) == "percentage"

    def test_currency(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        cell = ws.cell(row=1, column=1, value=99.99)
        cell.number_format = '"$"#,##0.00'
        assert detect_semantic_type(cell) == "currency"

    def test_date(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        cell = ws.cell(row=1, column=1, value=date(2024, 6, 15))
        cell.number_format = "yyyy-mm-dd"
        assert detect_semantic_type(cell) == "date"

    def test_email(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="user@domain.org")
        assert detect_semantic_type(ws.cell(row=1, column=1)) == "email"

    def test_text(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="hello world")
        assert detect_semantic_type(ws.cell(row=1, column=1)) == "text"

    def test_empty(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        assert detect_semantic_type(ws.cell(row=1, column=1)) == "empty"


class TestGetNumberFormatString:
    def test_general(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value=1)
        assert get_number_format_string(ws.cell(row=1, column=1)) == "General"

    def test_custom_format(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        cell = ws.cell(row=1, column=1, value=0.5)
        cell.number_format = "0.00%"
        assert get_number_format_string(cell) == "0.00%"


class TestStyleFingerprint:
    def test_same_style_same_fingerprint(self, tmp_xlsx):
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="a")
        ws.cell(row=1, column=2, value="b")
        fp1 = get_cell_style_fingerprint(ws.cell(row=1, column=1))
        fp2 = get_cell_style_fingerprint(ws.cell(row=1, column=2))
        # Default style cells should have same fingerprint
        assert fp1 == fp2

    def test_different_style_different_fingerprint(self, tmp_xlsx):
        from openpyxl.styles import Font
        wb, ws, _ = tmp_xlsx
        ws.cell(row=1, column=1, value="a")
        ws.cell(row=1, column=2, value="b").font = Font(bold=True)
        fp1 = get_cell_style_fingerprint(ws.cell(row=1, column=1))
        fp2 = get_cell_style_fingerprint(ws.cell(row=1, column=2))
        assert fp1 != fp2


class TestHelpers:
    def test_cell_coord(self):
        assert cell_coord(1, 1) == "A1"
        assert cell_coord(5, 3) == "C5"
        assert cell_coord(100, 27) == "AA100"

    def test_split_cell_ref(self):
        assert split_cell_ref("A1") == ("A", 1)
        assert split_cell_ref("AB123") == ("AB", 123)
        assert split_cell_ref("Z99") == ("Z", 99)
