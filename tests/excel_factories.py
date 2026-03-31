"""Factory functions to create test Excel files of varying complexity.

Each factory returns the path to the created .xlsx file.
All files are created in a temp directory and cleaned up by pytest fixtures.
"""
from __future__ import annotations

import os
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(bottom=Side(style="thin"))
GREEN_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")


def make_empty_sheet(path: str) -> str:
    """Completely empty workbook — single sheet, no data."""
    wb = openpyxl.Workbook()
    wb.save(path)
    return path


def make_single_cell(path: str) -> str:
    """One cell with a value."""
    wb = openpyxl.Workbook()
    wb.active.cell(row=1, column=1, value="hello")
    wb.save(path)
    return path


def make_tiny_with_formulas(path: str) -> str:
    """5 rows, 4 cols, basic SUM formulas. Tests formula preservation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    for c, h in enumerate(["Product", "Q1", "Q2", "Total"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = BOLD
        cell.alignment = CENTER

    data = [
        ("Widget A", 100, 200, "=B2+C2"),
        ("Widget B", 150, 300, "=B3+C3"),
        ("Total", "=SUM(B2:B3)", "=SUM(C2:C3)", "=SUM(D2:D3)"),
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)
    return path


def make_adjacent_tables_no_gap(path: str) -> str:
    """Two tables stacked vertically with NO empty row between them.
    Table 1: Sales (rows 1-5), Table 2: Employees (rows 6-10).
    Different header styles so boundary detection should catch the transition.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1
    for c, h in enumerate(["Region", "Revenue", "Cost"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for r in range(2, 6):
        ws.cell(row=r, column=1, value=f"Region {r-1}")
        ws.cell(row=r, column=2, value=(r - 1) * 10000)
        ws.cell(row=r, column=3, value=(r - 1) * 6000)

    # Table 2 — immediately after, no gap
    for c, h in enumerate(["Name", "Dept", "Salary"], 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = GREEN_FILL

    employees = [("Alice", "Eng", 90000), ("Bob", "Sales", 80000),
                 ("Carol", "Eng", 95000), ("Dave", "HR", 75000)]
    for i, emp in enumerate(employees):
        for c, val in enumerate(emp, 1):
            ws.cell(row=7 + i, column=c, value=val)

    wb.save(path)
    return path


def make_adjacent_tables_same_style(path: str) -> str:
    """Two tables stacked with NO gap and SAME header styling.
    Must rely on data-type transition detection (numeric→text).
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1: numeric data (rows 1-5)
    for c, h in enumerate(["ID", "Value", "Score"], 1):
        ws.cell(row=1, column=c, value=h).font = BOLD

    for r in range(2, 6):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=(r - 1) * 3.14)
        ws.cell(row=r, column=3, value=(r - 1) * 10)

    # Table 2: text-heavy (rows 6-10), same bold header style
    for c, h in enumerate(["Category", "Label", "Notes"], 1):
        ws.cell(row=6, column=c, value=h).font = BOLD

    for r in range(7, 11):
        ws.cell(row=r, column=1, value=f"Cat-{r-6}")
        ws.cell(row=r, column=2, value=f"Label-{r-6}")
        ws.cell(row=r, column=3, value=f"Some note about item {r-6}")

    wb.save(path)
    return path


def make_merged_header(path: str) -> str:
    """Sheet with a merged title spanning A1:D1, then a table below."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="Annual Report 2024").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = CENTER

    for c, h in enumerate(["Quarter", "Revenue", "Expenses", "Profit"], 1):
        ws.cell(row=3, column=c, value=h).font = BOLD

    for r in range(4, 8):
        ws.cell(row=r, column=1, value=f"Q{r-3}")
        ws.cell(row=r, column=2, value=(r - 3) * 50000)
        ws.cell(row=r, column=3, value=(r - 3) * 30000)
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")

    wb.save(path)
    return path


def make_complex_formulas(path: str) -> str:
    """Sheet with various complex Excel formulas: VLOOKUP, IF, COUNTIF, etc."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = ["Name", "Score", "Grade", "Pass", "Rank"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = BOLD

    names = ["Alice", "Bob", "Carol", "Dave", "Eve"]
    scores = [92, 78, 85, 65, 95]
    for i, (name, score) in enumerate(zip(names, scores)):
        r = i + 2
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=score)
        ws.cell(row=r, column=3, value=f'=IF(B{r}>=90,"A",IF(B{r}>=80,"B",IF(B{r}>=70,"C","F")))')
        ws.cell(row=r, column=4, value=f'=IF(B{r}>=70,"Yes","No")')
        ws.cell(row=r, column=5, value=f"=RANK(B{r},$B$2:$B$6)")

    # Summary row
    r = 8
    ws.cell(row=r, column=1, value="Stats").font = BOLD
    ws.cell(row=r, column=2, value="=AVERAGE(B2:B6)")
    ws.cell(row=9, column=1, value="Pass Count")
    ws.cell(row=9, column=2, value='=COUNTIF(D2:D6,"Yes")')
    ws.cell(row=10, column=1, value="Max Score")
    ws.cell(row=10, column=2, value="=MAX(B2:B6)")
    ws.cell(row=11, column=1, value="Min Score")
    ws.cell(row=11, column=2, value="=MIN(B2:B6)")

    wb.save(path)
    return path


def make_mixed_types(path: str) -> str:
    """Sheet with all 9 semantic types: year, int, float, pct, scientific,
    date, time, currency, email."""
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["Type", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = BOLD

    test_data = [
        ("Year", 2024),
        ("Integer", 42),
        ("Float", 3.14159),
        ("Percentage", 0.856),
        ("Scientific", 6.022e23),
        ("Date", date(2024, 6, 15)),
        ("Time", datetime(2024, 1, 1, 14, 30, 0)),
        ("Currency", 99.99),
        ("Email", "test@example.com"),
        ("Text", "Hello World"),
        ("Boolean", True),
        ("Empty", None),
    ]

    for i, (label, val) in enumerate(test_data):
        r = i + 2
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=val)

        # Apply appropriate number formats
        if label == "Percentage":
            cell.number_format = "0.00%"
        elif label == "Scientific":
            cell.number_format = "0.00E+00"
        elif label == "Date":
            cell.number_format = "yyyy-mm-dd"
        elif label == "Time":
            cell.number_format = "hh:mm:ss"
        elif label == "Currency":
            cell.number_format = '"$"#,##0.00'
        elif label == "Year":
            cell.number_format = "yyyy"

    wb.save(path)
    return path


def make_medium_500_rows(path: str) -> str:
    """500 rows, 8 cols, formulas in every row, multiple data types."""
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ["ID", "Date", "Product", "Qty", "Price", "Subtotal", "Tax", "Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = BOLD

    products = ["Widget", "Gadget", "Gizmo", "Doohickey", "Thingamajig"]
    for r in range(2, 502):
        idx = r - 2
        ws.cell(row=r, column=1, value=f"ORD-{idx+1:04d}")
        ws.cell(row=r, column=2, value=date(2024, (idx % 12) + 1, (idx % 28) + 1)).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=3, value=products[idx % 5])
        ws.cell(row=r, column=4, value=(idx % 10) + 1)
        ws.cell(row=r, column=5, value=round(9.99 + (idx % 20) * 5, 2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7, value=f"=F{r}*0.08").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=8, value=f"=F{r}+G{r}").number_format = '"$"#,##0.00'

    # Totals
    ws.cell(row=502, column=1, value="TOTAL").font = BOLD
    ws.cell(row=502, column=6, value="=SUM(F2:F501)").number_format = '"$"#,##0.00'
    ws.cell(row=502, column=8, value="=SUM(H2:H501)").number_format = '"$"#,##0.00'

    wb.save(path)
    return path


def make_large_3_tables(path: str) -> str:
    """~3000 rows, 3 tables on one sheet: table1 (rows 1-1001),
    empty row, table2 (1003-1503), table3 (1504-2004, no gap from table2).
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Table 1: Sensor data
    for c, h in enumerate(["Sensor", "Reading", "Status"], 1):
        ws.cell(row=1, column=c, value=h).font = BOLD

    for r in range(2, 1002):
        idx = r - 2
        ws.cell(row=r, column=1, value=f"S-{idx % 5 + 1:02d}")
        ws.cell(row=r, column=2, value=round(20.0 + (idx % 50) * 0.3, 2))
        ws.cell(row=r, column=3, value="OK" if idx % 7 != 0 else "ALERT")

    # Row 1002 is empty (gap)

    # Table 2: Location summary
    for c, h in enumerate(["Location", "Avg", "Max"], 1):
        ws.cell(row=1003, column=c, value=h).font = BOLD
    for r in range(1004, 1504):
        idx = r - 1004
        ws.cell(row=r, column=1, value=f"Bldg-{idx % 5 + 1}")
        ws.cell(row=r, column=2, value=round(21.0 + (idx % 10) * 0.5, 1))
        ws.cell(row=r, column=3, value=round(25.0 + (idx % 15) * 0.3, 1))

    # Table 3: NO gap from table 2
    for c, h in enumerate(["Shift", "Operator", "Output"], 1):
        cell = ws.cell(row=1504, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = GREEN_FILL
    for r in range(1505, 2005):
        idx = r - 1505
        ws.cell(row=r, column=1, value=["Morning", "Afternoon", "Night"][idx % 3])
        ws.cell(row=r, column=2, value=["John", "Jane", "Bob"][idx % 3])
        ws.cell(row=r, column=3, value=50 + idx % 30)

    wb.save(path)
    return path


def make_messy_form_layout(path: str) -> str:
    """Sparse form-like layout: labels in column A, values in column C,
    scattered data, no obvious table structure."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Company Info").font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    form_fields = [
        (3, "Company Name", "Acme Corp"),
        (4, "Address", "123 Main St"),
        (5, "City", "Springfield"),
        (6, "State", "IL"),
        (7, "ZIP", "62701"),
        (9, "Contact Info", None),  # Section header
        (10, "Phone", "(555) 123-4567"),
        (11, "Email", "info@acme.com"),
        (12, "Website", "www.acme.com"),
        (14, "Financial", None),  # Section header
        (15, "Revenue", 1500000),
        (16, "Employees", 42),
        (17, "Founded", date(1995, 3, 10)),
    ]

    for r, label, value in form_fields:
        ws.cell(row=r, column=1, value=label).font = BOLD
        if value is not None:
            ws.cell(row=r, column=3, value=value)

    # Section headers are bold + larger
    for r in (9, 14):
        ws.cell(row=r, column=1).font = Font(bold=True, size=12)

    wb.save(path)
    return path


def make_multi_sheet_cross_ref(path: str) -> str:
    """3 sheets with cross-sheet formula references."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Input"
    for c, h in enumerate(["Item", "Price", "Qty"], 1):
        ws1.cell(row=1, column=c, value=h).font = BOLD
    items = [("Apples", 1.50, 10), ("Bananas", 0.75, 20), ("Cherries", 3.00, 5)]
    for i, (item, price, qty) in enumerate(items):
        r = i + 2
        ws1.cell(row=r, column=1, value=item)
        ws1.cell(row=r, column=2, value=price).number_format = '"$"#,##0.00'
        ws1.cell(row=r, column=3, value=qty)

    ws2 = wb.create_sheet("Calc")
    ws2.cell(row=1, column=1, value="Item").font = BOLD
    ws2.cell(row=1, column=2, value="Subtotal").font = BOLD
    for r in range(2, 5):
        ws2.cell(row=r, column=1, value=f"=Input!A{r}")
        ws2.cell(row=r, column=2, value=f"=Input!B{r}*Input!C{r}").number_format = '"$"#,##0.00'
    ws2.cell(row=5, column=1, value="Grand Total").font = BOLD
    ws2.cell(row=5, column=2, value="=SUM(B2:B4)").number_format = '"$"#,##0.00'

    ws3 = wb.create_sheet("Summary")
    ws3.merge_cells("A1:B1")
    ws3.cell(row=1, column=1, value="Order Summary").font = Font(bold=True, size=14)
    ws3.cell(row=3, column=1, value="Total Items")
    ws3.cell(row=3, column=2, value="=SUM(Input!C2:C4)")
    ws3.cell(row=4, column=1, value="Total Cost")
    ws3.cell(row=4, column=2, value="=Calc!B5")

    wb.save(path)
    return path
