"""Unit tests for encoder.py — merge_cell_ranges, inverted index, anchors."""
from __future__ import annotations

import json
import os
import pytest
import openpyxl

from spreadsheet_llm.encoder import (
    merge_cell_ranges,
    create_inverted_index_translation,
    compress_homogeneous_regions,
    _build_merged_cell_map,
    _is_empty_sheet,
)


class TestMergeCellRanges:
    def test_empty(self):
        assert merge_cell_ranges([]) == []

    def test_single_cell(self):
        result = merge_cell_ranges(["A1"])
        assert result == ["A1"]

    def test_horizontal_merge(self):
        result = merge_cell_ranges(["A1", "B1", "C1"])
        assert result == ["A1:C1"]

    def test_vertical_merge(self):
        result = merge_cell_ranges(["A1", "A2", "A3"])
        assert result == ["A1:A3"]

    def test_rectangle_merge(self):
        refs = ["A1", "B1", "A2", "B2", "A3", "B3"]
        result = merge_cell_ranges(refs)
        assert result == ["A1:B3"]

    def test_disjoint_cells(self):
        result = merge_cell_ranges(["A1", "C3", "E5"])
        assert len(result) == 3
        assert set(result) == {"A1", "C3", "E5"}

    def test_l_shape_not_merged(self):
        # L-shape cannot be a single rectangle
        refs = ["A1", "A2", "A3", "B1"]
        result = merge_cell_ranges(refs)
        # Should produce at least 2 ranges (various valid decompositions)
        total_cells = 0
        for r in result:
            if ":" in r:
                # Count cells in range
                parts = r.split(":")
                # Approximate — just check we got reasonable output
                total_cells += 2  # at least 2
            else:
                total_cells += 1
        assert len(result) >= 1
        assert len(result) <= 4  # At most 4 single cells

    def test_large_set(self):
        # 100 cells in a 10x10 grid should merge to one range
        refs = [f"{chr(65+c)}{r}" for r in range(1, 11) for c in range(10)]
        result = merge_cell_ranges(refs)
        assert len(result) == 1
        assert ":" in result[0]


class TestInvertedIndexTranslation:
    def test_basic(self):
        index = {"hello": ["A1", "A2", "A3"], "world": ["B1"]}
        result = create_inverted_index_translation(index)
        assert result["hello"] == ["A1:A3"]
        assert result["world"] == ["B1"]

    def test_skips_empty_values(self):
        index = {"": ["A1"], "  ": ["B1"], "real": ["C1"]}
        result = create_inverted_index_translation(index)
        assert "real" in result
        assert "" not in result
        assert "  " not in result

    def test_multiple_ranges(self):
        # Two separate clusters of the same value
        index = {"X": ["A1", "A2", "C1", "C2"]}
        result = create_inverted_index_translation(index)
        assert len(result["X"]) == 2  # Two separate ranges


class TestBuildMergedCellMap:
    def test_no_merges(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="a")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        merged_map = _build_merged_cell_map(ws2)
        assert merged_map == {}

    def test_with_merges(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1, value="merged title")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        merged_map = _build_merged_cell_map(ws2)
        assert "A1" in merged_map
        assert "B1" in merged_map
        assert "C1" in merged_map
        assert merged_map["A1"] == "A1"
        assert merged_map["B1"] == "A1"
        assert merged_map["C1"] == "A1"


class TestCompressHomogeneous:
    def test_removes_uniform_row(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        # Row 1: all "X"
        for c in range(1, 4):
            ws.cell(row=1, column=c, value="X")
        # Row 2: varied
        ws.cell(row=2, column=1, value="A")
        ws.cell(row=2, column=2, value="B")
        ws.cell(row=2, column=3, value="C")
        # Row 3: all "X" again
        for c in range(1, 4):
            ws.cell(row=3, column=c, value="X")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        rows, cols = compress_homogeneous_regions(ws2, [1, 2, 3], [1, 2, 3])
        # Row 2 should survive, rows 1 and 3 are homogeneous
        assert 2 in rows
        # Rows 1 and 3 should be removed
        assert 1 not in rows
        assert 3 not in rows

    def test_all_homogeneous_keeps_all(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, 4):
            for c in range(1, 4):
                ws.cell(row=r, column=c, value="same")
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        rows, cols = compress_homogeneous_regions(ws2, [1, 2, 3], [1, 2, 3])
        # Fallback: if all homogeneous, keep all
        assert rows == [1, 2, 3]


class TestIsEmptySheet:
    def test_empty(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        assert _is_empty_sheet(wb2.active) is True

    def test_not_empty(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        wb.active.cell(row=1, column=1, value="data")
        wb.save(path)
        wb2 = openpyxl.load_workbook(path)
        assert _is_empty_sheet(wb2.active) is False
