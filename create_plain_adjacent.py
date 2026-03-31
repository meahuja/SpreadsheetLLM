"""Create a plain Excel with 2 adjacent tables — NO formatting at all.
No bold, no colors, no borders. Just raw text and numbers."""
import json
import os
import sys
from datetime import date

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from spreadsheet_llm.encoder import encode_spreadsheet

path = os.path.join(os.path.dirname(__file__), "plain_adjacent.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# ------ TABLE 1: Fruit Prices (rows 1-6) ------
ws["A1"] = "Fruit"
ws["B1"] = "Price"
ws["C1"] = "Stock"
ws["D1"] = "Total Value"

ws["A2"] = "Apple"
ws["B2"] = 1.50
ws["C2"] = 200
ws["D2"] = "=B2*C2"

ws["A3"] = "Banana"
ws["B3"] = 0.75
ws["C3"] = 350
ws["D3"] = "=B3*C3"

ws["A4"] = "Cherry"
ws["B4"] = 3.00
ws["C4"] = 100
ws["D4"] = "=B4*C4"

ws["A5"] = "Mango"
ws["B5"] = 2.25
ws["C5"] = 150
ws["D5"] = "=B5*C5"

ws["A6"] = "Total"
ws["B6"] = "=AVERAGE(B2:B5)"
ws["C6"] = "=SUM(C2:C5)"
ws["D6"] = "=SUM(D2:D5)"

# ------ TABLE 2: Student Grades (rows 7-12) — NO GAP ------
ws["A7"] = "Student"
ws["B7"] = "Math"
ws["C7"] = "Science"
ws["D7"] = "Average"

ws["A8"] = "John"
ws["B8"] = 85
ws["C8"] = 92
ws["D8"] = "=AVERAGE(B8:C8)"

ws["A9"] = "Sara"
ws["B9"] = 78
ws["C9"] = 88
ws["D9"] = "=AVERAGE(B9:C9)"

ws["A10"] = "Mike"
ws["B10"] = 92
ws["C10"] = 76
ws["D10"] = "=AVERAGE(B10:C10)"

ws["A11"] = "Lisa"
ws["B11"] = 95
ws["C11"] = 98
ws["D11"] = "=AVERAGE(B11:C11)"

ws["A12"] = "Class Avg"
ws["B12"] = "=AVERAGE(B8:B11)"
ws["C12"] = "=AVERAGE(C8:C11)"
ws["D12"] = "=AVERAGE(D8:D11)"

wb.save(path)
print(f"Created: {path}")
print()

# Show raw content
print("RAW EXCEL CONTENT (no formatting at all):")
print("-" * 60)
wb2 = openpyxl.load_workbook(path, data_only=False)
ws2 = wb2.active
for r in range(1, 13):
    cells = []
    for c in range(1, 5):
        v = ws2.cell(row=r, column=c).value
        cells.append(str(v) if v is not None else "")
    tag = ""
    if r == 1: tag = "  <-- Table 1 header"
    if r == 6: tag = "  <-- Table 1 total"
    if r == 7: tag = "  <-- Table 2 header (NO GAP!)"
    if r == 12: tag = "  <-- Table 2 total"
    print(f"  Row {r:2d}: {' | '.join(f'{c:15s}' for c in cells)}{tag}")

# Encode
print()
print("=" * 60)
print("ENCODED JSON")
print("=" * 60)
result = encode_spreadsheet(path, k=2)
print(json.dumps(result, indent=2, ensure_ascii=False))
